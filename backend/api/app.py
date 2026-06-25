"""
Teacher Buddy - Flask REST API
Provides authentication, chat management, and AI integration endpoints
"""

from flask import Flask, request, jsonify
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt_identity
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta
import requests
import os
from sqlalchemy import func
import tempfile
import pdfplumber
from docx import Document
from openpyxl import load_workbook
import gzip
import base64

app = Flask(__name__)

# ============================================================================
# Configuration
# ============================================================================

# Security keys - override via environment variables in production
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')
app.config['JWT_SECRET_KEY'] = os.environ.get('JWT_SECRET_KEY', 'jwt-secret-key-change-in-production')

# JWT settings
app.config['JWT_ACCESS_TOKEN_EXPIRES'] = timedelta(hours=24)  # Tokens expire after 24 hours
app.config['JWT_TOKEN_LOCATION'] = ['headers']  # Look for token in Authorization header
app.config['JWT_HEADER_NAME'] = 'Authorization'
app.config['JWT_HEADER_TYPE'] = 'Bearer'  # Format: "Bearer <token>"
app.config['JWT_CSRF_CHECK_FORM'] = False  # Disable CSRF for API-only usage

# Database configuration
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get(
    'DATABASE_URL',
    'postgresql://teacherbuddy:teacherbuddy123@postgres:5432/teacherbuddy'
)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False  # Disable Flask-SQLAlchemy event system
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024  # 5MB upload limit

ALLOWED_RESOURCE_EXTENSIONS = {'.pdf', '.txt', '.docx', '.xlsx'}

# ============================================================================
# Initialize Extensions
# ============================================================================

# CORS - Allow frontend on both common Vite ports
CORS(app, origins=['http://localhost:5173', 'http://localhost:5174'])

# Database ORM
db = SQLAlchemy(app)

# JWT Manager for token handling
jwt = JWTManager(app)

# ============================================================================
# JWT Error Handlers
# ============================================================================
@jwt.expired_token_loader
def expired_token_callback(jwt_header, jwt_payload):
    return jsonify({'error': 'Token has expired'}), 401

@jwt.invalid_token_loader
def invalid_token_callback(error):
    return jsonify({'error': 'Invalid token'}), 422

@jwt.unauthorized_loader
def missing_token_callback(error):
    return jsonify({'error': 'Authorization token is missing'}), 401

# Database Models
class User(db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    is_admin = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    # Teacher preferences
    grade_level = db.Column(db.String(100), default='')
    curriculum = db.Column(db.String(100), default='')
    class_size = db.Column(db.String(50), default='')
    teaching_style = db.Column(db.String(100), default='')
    chat_sessions = db.relationship('ChatSession', backref='user', lazy=True, cascade='all, delete-orphan')
    resource_files = db.relationship('ResourceFile', backref='user', lazy=True, cascade='all, delete-orphan')

class ChatSession(db.Model):
    __tablename__ = 'chat_sessions'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    title = db.Column(db.String(200), default='New Chat')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    messages = db.relationship('Message', backref='session', lazy=True, cascade='all, delete-orphan')

class Message(db.Model):
    __tablename__ = 'messages'
    id = db.Column(db.Integer, primary_key=True)
    session_id = db.Column(db.Integer, db.ForeignKey('chat_sessions.id'), nullable=False)
    role = db.Column(db.String(20), nullable=False)  # 'user' or 'assistant'
    content = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class ResourceFile(db.Model):
    __tablename__ = 'resource_files'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    filename = db.Column(db.String(255), nullable=False)
    filetype = db.Column(db.String(20), nullable=False)
    content = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class LessonPlan(db.Model):
    __tablename__ = 'lesson_plans'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    title = db.Column(db.String(255), nullable=False)
    instructions = db.Column(db.Text, nullable=False)
    resource_file_id = db.Column(db.Integer, db.ForeignKey('resource_files.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class Note(db.Model):
    __tablename__ = 'notes'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    title = db.Column(db.String(255), nullable=False)
    content = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class CalendarEvent(db.Model):
    __tablename__ = 'calendar_events'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    date = db.Column(db.String(10), nullable=False)  # YYYY-MM-DD format
    title = db.Column(db.String(255), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class CalendarClass(db.Model):
    __tablename__ = 'calendar_classes'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    name = db.Column(db.String(255), nullable=False)
    date = db.Column(db.String(10), nullable=False)  # YYYY-MM-DD format
    time_from = db.Column(db.String(5), nullable=False)  # HH:MM format
    time_to = db.Column(db.String(5), nullable=False)    # HH:MM format
    room = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Assignment(db.Model):
    __tablename__ = 'assignments'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    title = db.Column(db.String(255), nullable=False)
    due_date = db.Column(db.String(10), nullable=False)  # YYYY-MM-DD format
    class_name = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

# Initialize database
with app.app_context():
    db.create_all()
    # Create default admin user if it doesn't exist
    admin = User.query.filter_by(username='admin').first()
    if not admin:
        admin = User(
            username='admin',
            email='admin@teacherbuddy.com',
            password_hash=generate_password_hash('admin123'),
            is_admin=True
        )
        db.session.add(admin)
        db.session.commit()
        print('Default admin user created (username: admin, password: admin123)')
    elif not admin.is_admin:
        auto_promote = os.environ.get('TEACHERBUDDY_AUTO_PROMOTE_ADMIN', '').strip().lower() in ('1', 'true', 'yes', 'on')
        if auto_promote:
            admin.is_admin = True
            db.session.commit()
            print('Default admin user promoted to admin (username: admin)')
        else:
            print(
                "Default 'admin' user exists but is not an admin. "
                "Set TEACHERBUDDY_AUTO_PROMOTE_ADMIN=true to auto-promote on startup (dev only), "
                "or reset your database volume."
            )

# Ollama Configuration
OLLAMA_URL = os.environ.get('OLLAMA_URL', 'http://ollama:11434/api/generate')
OLLAMA_MODEL = os.environ.get('OLLAMA_MODEL', 'llama3')

# Helper function to call Ollama
def call_ollama(prompt, system_prompt=None):
    """Call Ollama API and return the response"""
    try:
        full_prompt = prompt
        if system_prompt:
            full_prompt = f"{system_prompt}\n\nUser: {prompt}"
        
        payload = {
            "model": OLLAMA_MODEL,
            "prompt": full_prompt,
            "stream": False
        }
        
        # Increased timeout to 300 seconds (5 minutes) for slower computers
        response = requests.post(OLLAMA_URL, json=payload, timeout=300)
        response.raise_for_status()
        
        result = response.json()
        return result.get('response', 'No response from AI')
    except requests.exceptions.RequestException as e:
        print(f"Error calling Ollama: {e}")
        return f"Error: Unable to get AI response. {str(e)}"

def is_allowed_resource(filename: str) -> bool:
    _, ext = os.path.splitext(filename.lower())
    return ext in ALLOWED_RESOURCE_EXTENSIONS

def compress_text(text: str) -> str:
    """Compress text using gzip and encode to base64 for storage"""
    compressed = gzip.compress(text.encode('utf-8'))
    return base64.b64encode(compressed).decode('ascii')

def decompress_text(compressed: str) -> str:
    """Decompress base64-encoded gzipped text"""
    try:
        decoded = base64.b64decode(compressed.encode('ascii'))
        decompressed = gzip.decompress(decoded)
        return decompressed.decode('utf-8')
    except Exception:
        # If decompression fails, assume it's uncompressed text (backward compatibility)
        return compressed

def extract_text_from_file(filename: str, file_storage) -> str:
    _, ext = os.path.splitext(filename.lower())

    if ext == '.txt':
        return file_storage.read().decode('utf-8', errors='ignore')

    if ext == '.docx':
        document = Document(file_storage)
        return "\n".join([para.text for para in document.paragraphs])

    if ext == '.xlsx':
        workbook = load_workbook(file_storage, data_only=True)
        text_lines = []
        for sheet in workbook.worksheets:
            text_lines.append(f"Sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                row_values = [str(cell) for cell in row if cell is not None]
                if row_values:
                    text_lines.append("\t".join(row_values))
        return "\n".join(text_lines)

    if ext == '.pdf':
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=True) as temp_file:
            file_storage.save(temp_file.name)
            text_chunks = []
            with pdfplumber.open(temp_file.name) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text() or ''
                    if page_text:
                        text_chunks.append(page_text)
            return "\n".join(text_chunks)

    return ''

# Routes

@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({'status': 'healthy', 'service': 'Teacher Buddy API'}), 200

# Authentication Routes
@app.route('/api/auth/register', methods=['POST'])
def register():
    """Register a new user"""
    try:
        data = request.get_json()
        
        # Validate required fields
        if not data.get('username') or not data.get('email') or not data.get('password'):
            return jsonify({'error': 'Username, email, and password are required'}), 400
        
        # Check if user already exists
        if User.query.filter_by(username=data['username']).first():
            return jsonify({'error': 'Username already exists'}), 409
        
        if User.query.filter_by(email=data['email']).first():
            return jsonify({'error': 'Email already exists'}), 409
        
        requested_is_admin = bool(data.get('isAdmin', False))
        # Only allow self-service admin registration in debug mode unless explicitly enabled.
        allow_admin_signup = app.debug or os.environ.get('TEACHERBUDDY_ALLOW_ADMIN_SIGNUP', '').lower() in ('1', 'true', 'yes')
        if requested_is_admin and not allow_admin_signup:
            return jsonify({'error': 'Admin registration is disabled'}), 403

        # Create new user
        user = User(
            username=data['username'],
            email=data['email'],
            password_hash=generate_password_hash(data['password']),
            is_admin=(requested_is_admin and allow_admin_signup)
        )
        
        db.session.add(user)
        db.session.commit()
        
        # Create access token - identity must be a string
        access_token = create_access_token(identity=str(user.id))
        
        return jsonify({
            'message': 'User registered successfully',
            'token': access_token,
            'user': {
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'isAdmin': user.is_admin
            }
        }), 201
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/auth/login', methods=['POST'])
def login():
    """Login user and return JWT token"""
    try:
        data = request.get_json()
        
        # Validate required fields
        if not data.get('username') or not data.get('password'):
            return jsonify({'error': 'Username and password are required'}), 400
        
        # Find user
        user = User.query.filter_by(username=data['username']).first()
        
        if not user or not check_password_hash(user.password_hash, data['password']):
            return jsonify({'error': 'Invalid username or password'}), 401
        
        # Create access token - identity must be a string
        access_token = create_access_token(identity=str(user.id))
        
        return jsonify({
            'message': 'Login successful',
            'token': access_token,
            'user': {
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'isAdmin': user.is_admin
            }
        }), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/auth/verify', methods=['GET'])
@jwt_required()
def verify_token():
    """Verify JWT token and return user info"""
    try:
        current_user_id = int(get_jwt_identity())
        user = User.query.get(current_user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        return jsonify({
            'user': {
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'isAdmin': user.is_admin
            }
        }), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Chat Routes
@app.route('/api/chat/sessions', methods=['GET'])
@jwt_required()
def get_chat_sessions():
    """Get all chat sessions for the current user"""
    try:
        current_user_id = int(get_jwt_identity())
        sessions = ChatSession.query.filter_by(user_id=current_user_id).order_by(ChatSession.created_at.desc()).all()
        
        return jsonify({
            'sessions': [{
                'id': session.id,
                'title': session.title,
                'created_at': session.created_at.isoformat(),
                'message_count': len(session.messages)
            } for session in sessions]
        }), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/chat/sessions/<int:session_id>', methods=['GET'])
@jwt_required()
def get_chat_history(session_id):
    """Get chat history for a specific session"""
    try:
        current_user_id = int(get_jwt_identity())
        session = ChatSession.query.filter_by(id=session_id, user_id=current_user_id).first()
        
        if not session:
            return jsonify({'error': 'Chat session not found'}), 404
        
        messages = Message.query.filter_by(session_id=session_id).order_by(Message.created_at).all()
        
        return jsonify({
            'session': {
                'id': session.id,
                'title': session.title,
                'created_at': session.created_at.isoformat()
            },
            'messages': [{
                'id': msg.id,
                'role': msg.role,
                'content': msg.content,
                'created_at': msg.created_at.isoformat()
            } for msg in messages]
        }), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/chat/sessions/<int:session_id>', methods=['DELETE'])
@jwt_required()
def delete_chat_session(session_id):
    """Delete a chat session and all its messages"""
    try:
        current_user_id = int(get_jwt_identity())
        session = ChatSession.query.filter_by(id=session_id, user_id=current_user_id).first()
        
        if not session:
            return jsonify({'error': 'Chat session not found'}), 404
        
        db.session.delete(session)
        db.session.commit()
        
        return jsonify({'message': 'Chat session deleted successfully'}), 200
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/chat/send', methods=['POST'])
@jwt_required()
def send_message():
    """Send a message and get AI response"""
    try:
        current_user_id = int(get_jwt_identity())
        data = request.get_json()
        
        if not data.get('message'):
            return jsonify({'error': 'Message is required'}), 400
        
        session_id = data.get('session_id')
        
        # Create new session if not provided
        if not session_id:
            session = ChatSession(
                user_id=current_user_id,
                title=data['message'][:50] + ('...' if len(data['message']) > 50 else '')
            )
            db.session.add(session)
            db.session.commit()
            session_id = session.id
        else:
            # Verify session belongs to user
            session = ChatSession.query.filter_by(id=session_id, user_id=current_user_id).first()
            if not session:
                return jsonify({'error': 'Chat session not found'}), 404
        
        # Save user message
        user_message = Message(
            session_id=session_id,
            role='user',
            content=data['message']
        )
        db.session.add(user_message)
        db.session.commit()
        
        # Get conversation context
        recent_messages = Message.query.filter_by(session_id=session_id).order_by(Message.created_at.desc()).limit(10).all()
        recent_messages.reverse()
        
        # Build context for Ollama
        context = ""
        for msg in recent_messages[:-1]:  # Exclude the message we just added
            context += f"{msg.role.capitalize()}: {msg.content}\n"

        # Resource context (selected file preferred, otherwise latest 3)
        resource_context = ""
        selected_resource_id = data.get('resource_id')
        if selected_resource_id:
            selected_resource = ResourceFile.query.filter_by(
                id=selected_resource_id,
                user_id=current_user_id
            ).first()
            if selected_resource:
                decompressed_content = decompress_text(selected_resource.content or '')
                snippet = decompressed_content[:4000]
                if snippet:
                    resource_context += f"\n[Resource: {selected_resource.filename}]\n{snippet}\n"
        else:
            resources = (
                ResourceFile.query.filter_by(user_id=current_user_id)
                .order_by(ResourceFile.created_at.desc())
                .limit(3)
                .all()
            )
            for resource in resources:
                decompressed_content = decompress_text(resource.content or '')
                snippet = decompressed_content[:4000]
                if snippet:
                    resource_context += f"\n[Resource: {resource.filename}]\n{snippet}\n"

        # Optional lesson plan context
        lesson_context = data.get('context') or ''
        if lesson_context:
            lesson_context = f"\n[Lesson Plan Context]\n{lesson_context}\n"
        
        # System prompt for teacher assistant
        system_prompt = """You are an AI teaching assistant helping teachers with their work. You can:
- Explain complex topics in simple terms
- Generate quiz questions and practice problems
- Simplify educational material
- Solve math problems
- Provide teaching strategies and tips

Be helpful, concise, and professional."""
        
        # Get AI response
        full_prompt = context + lesson_context + resource_context + f"\nUser: {data['message']}\nAssistant:"
        ai_response = call_ollama(full_prompt, system_prompt)
        
        # Save AI message
        ai_message = Message(
            session_id=session_id,
            role='assistant',
            content=ai_response
        )
        db.session.add(ai_message)
        db.session.commit()
        
        return jsonify({
            'session_id': session_id,
            'user_message': {
                'id': user_message.id,
                'role': 'user',
                'content': user_message.content,
                'created_at': user_message.created_at.isoformat()
            },
            'ai_message': {
                'id': ai_message.id,
                'role': 'assistant',
                'content': ai_message.content,
                'created_at': ai_message.created_at.isoformat()
            }
        }), 200
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

# Resource Routes
@app.route('/api/resources', methods=['GET'])
@jwt_required()
def list_resources():
    """List uploaded resource files for the current user"""
    try:
        current_user_id = int(get_jwt_identity())
        resources = (
            ResourceFile.query.filter_by(user_id=current_user_id)
            .order_by(ResourceFile.created_at.desc())
            .all()
        )
        return jsonify({
            'resources': [{
                'id': r.id,
                'filename': r.filename,
                'filetype': r.filetype,
                'created_at': r.created_at.isoformat()
            } for r in resources]
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/resources/upload', methods=['POST'])
@jwt_required()
def upload_resource():
    """Upload a resource file and extract text for AI context"""
    try:
        current_user_id = int(get_jwt_identity())
        if 'file' not in request.files:
            return jsonify({'error': 'File is required'}), 400

        file = request.files['file']
        if not file or file.filename == '':
            return jsonify({'error': 'File is required'}), 400

        filename = secure_filename(file.filename)
        if not is_allowed_resource(filename):
            return jsonify({'error': 'Unsupported file type'}), 400

        extracted_text = extract_text_from_file(filename, file)
        if not extracted_text.strip():
            return jsonify({'error': 'Could not extract text from file'}), 400

        # Truncate and compress the content
        truncated_text = extracted_text[:20000]
        compressed_content = compress_text(truncated_text)
        _, ext = os.path.splitext(filename.lower())

        resource = ResourceFile(
            user_id=current_user_id,
            filename=filename,
            filetype=ext.replace('.', ''),
            content=compressed_content
        )
        db.session.add(resource)
        db.session.commit()

        return jsonify({
            'id': resource.id,
            'filename': resource.filename,
            'filetype': resource.filetype,
            'created_at': resource.created_at.isoformat(),
            'content_length': len(truncated_text),
            'compressed_size': len(compressed_content)
        }), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

# Admin Routes
@app.route('/api/admin/stats', methods=['GET'])
@jwt_required()
def get_admin_stats():
    """Get admin statistics"""
    try:
        current_user_id = int(get_jwt_identity())
        user = User.query.get(current_user_id)
        
        if not user or not user.is_admin:
            return jsonify({'error': 'Unauthorized'}), 403
        
        # Get statistics
        total_users = User.query.count()
        total_sessions = ChatSession.query.count()
        total_messages = Message.query.count()
        
        # Get recent activity (last 7 days)
        seven_days_ago = datetime.utcnow() - timedelta(days=7)
        recent_users = User.query.filter(User.created_at >= seven_days_ago).count()
        recent_sessions = ChatSession.query.filter(ChatSession.created_at >= seven_days_ago).count()
        
        # Get active users (users with messages in last 7 days)
        active_users = db.session.query(func.count(func.distinct(ChatSession.user_id)))\
            .join(Message, ChatSession.id == Message.session_id)\
            .filter(Message.created_at >= seven_days_ago)\
            .scalar()
        
        return jsonify({
            'total_users': total_users,
            'total_chats': total_sessions,
            'total_messages': total_messages,
            'active_users': active_users or 0,
            'new_users_this_week': recent_users,
            'new_chats_this_week': recent_sessions,
            'avg_messages_per_chat': round(total_messages / total_sessions, 2) if total_sessions > 0 else 0
        }), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================================
# Lesson Plans Endpoints
# ============================================================================

@app.route('/api/lesson-plans', methods=['GET'])
@jwt_required()
def get_lesson_plans():
    """Get all lesson plans for the current user"""
    user_id = get_jwt_identity()
    plans = LessonPlan.query.filter_by(user_id=user_id).order_by(LessonPlan.created_at.desc()).all()
    return jsonify([{
        'id': plan.id,
        'title': plan.title,
        'instructions': plan.instructions,
        'resourceFileId': plan.resource_file_id,
        'createdAt': plan.created_at.isoformat(),
        'updatedAt': plan.updated_at.isoformat()
    } for plan in plans]), 200

@app.route('/api/lesson-plans', methods=['POST'])
@jwt_required()
def create_lesson_plan():
    """Create a new lesson plan"""
    user_id = get_jwt_identity()
    data = request.get_json()
    
    if not data.get('title') or not data.get('instructions'):
        return jsonify({'error': 'Title and instructions are required'}), 400
    
    plan = LessonPlan(
        user_id=user_id,
        title=data['title'],
        instructions=data['instructions'],
        resource_file_id=data.get('resourceFileId')
    )
    db.session.add(plan)
    db.session.commit()
    
    return jsonify({
        'id': plan.id,
        'title': plan.title,
        'instructions': plan.instructions,
        'resourceFileId': plan.resource_file_id,
        'createdAt': plan.created_at.isoformat()
    }), 201

@app.route('/api/lesson-plans/<int:plan_id>', methods=['DELETE'])
@jwt_required()
def delete_lesson_plan(plan_id):
    """Delete a lesson plan"""
    user_id = get_jwt_identity()
    plan = LessonPlan.query.filter_by(id=plan_id, user_id=user_id).first()
    
    if not plan:
        return jsonify({'error': 'Lesson plan not found'}), 404
    
    db.session.delete(plan)
    db.session.commit()
    return jsonify({'message': 'Lesson plan deleted'}), 200

# ============================================================================
# Notes Endpoints
# ============================================================================

@app.route('/api/notes', methods=['GET'])
@jwt_required()
def get_notes():
    """Get all notes for the current user"""
    user_id = get_jwt_identity()
    notes = Note.query.filter_by(user_id=user_id).order_by(Note.updated_at.desc()).all()
    return jsonify([{
        'id': note.id,
        'title': note.title,
        'content': note.content,
        'updatedAt': note.updated_at.isoformat()
    } for note in notes]), 200

@app.route('/api/notes', methods=['POST'])
@jwt_required()
def create_note():
    """Create a new note"""
    user_id = get_jwt_identity()
    data = request.get_json()
    
    note = Note(
        user_id=user_id,
        title=data.get('title', 'Untitled Note'),
        content=data.get('content', '')
    )
    db.session.add(note)
    db.session.commit()
    
    return jsonify({
        'id': note.id,
        'title': note.title,
        'content': note.content,
        'updatedAt': note.updated_at.isoformat()
    }), 201

@app.route('/api/notes/<int:note_id>', methods=['DELETE'])
@jwt_required()
def delete_note(note_id):
    """Delete a note"""
    user_id = get_jwt_identity()
    note = Note.query.filter_by(id=note_id, user_id=user_id).first()
    
    if not note:
        return jsonify({'error': 'Note not found'}), 404
    
    db.session.delete(note)
    db.session.commit()
    return jsonify({'message': 'Note deleted'}), 200

# ============================================================================
# Calendar Events Endpoints
# ============================================================================

@app.route('/api/calendar/events', methods=['GET'])
@jwt_required()
def get_calendar_events():
    """Get all calendar events for the current user"""
    user_id = get_jwt_identity()
    events = CalendarEvent.query.filter_by(user_id=user_id).all()
    return jsonify([{
        'date': event.date,
        'title': event.title
    } for event in events]), 200

@app.route('/api/calendar/events', methods=['POST'])
@jwt_required()
def create_calendar_event():
    """Create a calendar event"""
    user_id = get_jwt_identity()
    data = request.get_json()
    
    if not data.get('date') or not data.get('title'):
        return jsonify({'error': 'Date and title are required'}), 400
    
    event = CalendarEvent(
        user_id=user_id,
        date=data['date'],
        title=data['title']
    )
    db.session.add(event)
    db.session.commit()
    return jsonify({'date': event.date, 'title': event.title}), 201

@app.route('/api/calendar/events/<int:event_id>', methods=['DELETE'])
@jwt_required()
def delete_calendar_event(event_id):
    """Delete a calendar event"""
    user_id = get_jwt_identity()
    event = CalendarEvent.query.filter_by(id=event_id, user_id=user_id).first()
    
    if not event:
        return jsonify({'error': 'Event not found'}), 404
    
    db.session.delete(event)
    db.session.commit()
    return jsonify({'message': 'Event deleted'}), 200

# ============================================================================
# Calendar Classes Endpoints
# ============================================================================

@app.route('/api/calendar/classes', methods=['GET'])
@jwt_required()
def get_calendar_classes():
    """Get all calendar classes for the current user"""
    user_id = get_jwt_identity()
    classes = CalendarClass.query.filter_by(user_id=user_id).all()
    return jsonify([{
        'name': cls.name,
        'date': cls.date,
        'timeFrom': cls.time_from,
        'timeTo': cls.time_to,
        'room': cls.room
    } for cls in classes]), 200

@app.route('/api/calendar/classes', methods=['POST'])
@jwt_required()
def create_calendar_class():
    """Create a calendar class"""
    user_id = get_jwt_identity()
    data = request.get_json()
    
    if not data.get('date') or not data.get('name'):
        return jsonify({'error': 'Date and name are required'}), 400
    
    cls = CalendarClass(
        user_id=user_id,
        name=data['name'],
        date=data['date'],
        time_from=data.get('timeFrom', ''),
        time_to=data.get('timeTo', ''),
        room=data.get('room')
    )
    db.session.add(cls)
    db.session.commit()
    return jsonify({
        'name': cls.name,
        'date': cls.date,
        'timeFrom': cls.time_from,
        'timeTo': cls.time_to,
        'room': cls.room
    }), 201

@app.route('/api/calendar/classes/<int:class_id>', methods=['DELETE'])
@jwt_required()
def delete_calendar_class(class_id):
    """Delete a calendar class"""
    user_id = get_jwt_identity()
    cls = CalendarClass.query.filter_by(id=class_id, user_id=user_id).first()
    
    if not cls:
        return jsonify({'error': 'Class not found'}), 404
    
    db.session.delete(cls)
    db.session.commit()
    return jsonify({'message': 'Class deleted'}), 200

# ============================================================================
# Assignments Endpoints
# ============================================================================

@app.route('/api/assignments', methods=['GET'])
@jwt_required()
def get_assignments():
    """Get all assignments for the current user"""
    user_id = get_jwt_identity()
    assignments = Assignment.query.filter_by(user_id=user_id).all()
    return jsonify([{
        'title': assignment.title,
        'dueDate': assignment.due_date,
        'className': assignment.class_name
    } for assignment in assignments]), 200

@app.route('/api/assignments', methods=['POST'])
@jwt_required()
def create_assignment():
    """Create an assignment"""
    user_id = get_jwt_identity()
    data = request.get_json()
    
    if not data.get('title') or not data.get('dueDate'):
        return jsonify({'error': 'Title and due date are required'}), 400
    
    assignment = Assignment(
        user_id=user_id,
        title=data['title'],
        due_date=data['dueDate'],
        class_name=data.get('className')
    )
    db.session.add(assignment)
    db.session.commit()
    return jsonify({
        'title': assignment.title,
        'dueDate': assignment.due_date,
        'className': assignment.class_name
    }), 201

@app.route('/api/assignments/<int:assignment_id>', methods=['DELETE'])
@jwt_required()
def delete_assignment(assignment_id):
    """Delete an assignment"""
    user_id = get_jwt_identity()
    assignment = Assignment.query.filter_by(id=assignment_id, user_id=user_id).first()
    
    if not assignment:
        return jsonify({'error': 'Assignment not found'}), 404
    
    db.session.delete(assignment)
    db.session.commit()
    return jsonify({'message': 'Assignment deleted'}), 200

# ============================================================================
# Teacher Preferences Endpoints
# ============================================================================

@app.route('/api/preferences', methods=['GET'])
@jwt_required()
def get_preferences():
    """Get user's teacher preferences"""
    user_id = get_jwt_identity()
    user = User.query.get(user_id)
    
    if not user:
        return jsonify({'error': 'User not found'}), 404
    
    return jsonify({
        'gradeLevel': user.grade_level or '',
        'curriculum': user.curriculum or '',
        'classSize': user.class_size or '',
        'teachingStyle': user.teaching_style or ''
    }), 200

@app.route('/api/preferences', methods=['PUT'])
@jwt_required()
def update_preferences():
    """Update user's teacher preferences"""
    user_id = get_jwt_identity()
    user = User.query.get(user_id)
    
    if not user:
        return jsonify({'error': 'User not found'}), 404
    
    data = request.get_json()
    
    user.grade_level = data.get('gradeLevel', '')
    user.curriculum = data.get('curriculum', '')
    user.class_size = data.get('classSize', '')
    user.teaching_style = data.get('teachingStyle', '')
    
    db.session.commit()
    
    return jsonify({
        'gradeLevel': user.grade_level,
        'curriculum': user.curriculum,
        'classSize': user.class_size,
        'teachingStyle': user.teaching_style
    }), 200
    return jsonify({'message': 'Assignment deleted'}), 200

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
